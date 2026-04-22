"""Формирование Excel-отчётов для админ-панели.

Модуль строит три отчёта с помощью ``openpyxl``:

* ``build_events_report`` — детальная статистика по мероприятиям
  (для каждого мероприятия — сводные показатели по регистрациям).
* ``build_registrations_report`` — построчный список всех заявок.
* ``build_summary_report`` — «универсальный» отчёт с несколькими
  листами: сводка, разрезы по направлениям / типам / статусам.

Все функции принимают QuerySet-ы, уже подготовленные во вью,
и возвращают объект ``openpyxl.Workbook``. Ответственность за
конвертацию в HTTP-ответ остаётся на view-слое.
"""
from __future__ import annotations

from collections import Counter
from typing import Iterable

from django.db.models import Count, Q
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from catalog.models import Direction, Event, EventRegistration, EventType


HEADER_FONT = Font(bold=True, color='FFFFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='FF1F3A8A')
TITLE_FONT = Font(bold=True, size=14, color='FF111827')
SUBTITLE_FONT = Font(italic=True, size=10, color='FF6B7280')
TOTALS_FONT = Font(bold=True, size=11)
TOTALS_FILL = PatternFill('solid', fgColor='FFEFF6FF')


def _russian_now() -> str:
    """Возвращает текущую дату/время в локальном формате для подписи отчёта."""
    return timezone.localtime().strftime('%d.%m.%Y %H:%M')


def _format_datetime(value) -> str:
    """Приводит datetime к строке ``dd.mm.YYYY HH:MM`` с учётом TZ."""
    if not value:
        return ''
    if timezone.is_aware(value):
        value = timezone.localtime(value)
    return value.strftime('%d.%m.%Y %H:%M')


def _apply_header(row, ws: Worksheet) -> None:
    """Форматирует строку как заголовок таблицы (фон + жирный шрифт)."""
    for cell in row:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )


def _autosize_columns(ws: Worksheet, start_row: int = 1, max_width: int = 55) -> None:
    """Подбирает ширину колонок по содержимому (с верхней границей)."""
    if ws.max_row < start_row:
        return
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        longest = 0
        for row in ws.iter_rows(
            min_row=start_row, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
        ):
            for cell in row:
                value = '' if cell.value is None else str(cell.value)
                length = max(len(line) for line in value.splitlines() or [''])
                longest = max(longest, length)
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)


def _write_title(ws: Worksheet, title: str, subtitle: str = '') -> int:
    """Пишет заголовок + подзаголовок и возвращает первый «рабочий» номер строки."""
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    if subtitle:
        ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
        return 4
    return 3


def _annotate_event_stats(queryset) -> Iterable[Event]:
    """Добавляет счётчики регистраций на каждый из статусов к Event-у.

    После ``.annotate(...)`` у каждой записи появляются поля
    ``pending_total``, ``confirmed_total``, ``waitlist_total``,
    ``cancelled_total``, ``attended_total``, ``no_show_total``,
    ``registrations_total``.
    """
    Status = EventRegistration.Status
    return queryset.annotate(
        pending_total=Count('registrations', filter=Q(registrations__status=Status.PENDING)),
        confirmed_total=Count('registrations', filter=Q(registrations__status=Status.CONFIRMED)),
        waitlist_total=Count('registrations', filter=Q(registrations__status=Status.WAITLIST)),
        cancelled_total=Count('registrations', filter=Q(registrations__status=Status.CANCELLED)),
        attended_total=Count('registrations', filter=Q(registrations__status=Status.ATTENDED)),
        no_show_total=Count('registrations', filter=Q(registrations__status=Status.NO_SHOW)),
        registrations_total=Count('registrations'),
    )


# ---------------------------------------------------------------------------
# Отчёт 1: Мероприятия
# ---------------------------------------------------------------------------

EVENTS_COLUMNS = [
    ('ID', 8),
    ('Название', 40),
    ('Направление', 24),
    ('Тип', 20),
    ('Формат', 14),
    ('Статус', 16),
    ('Начало', 18),
    ('Окончание', 18),
    ('Место', 28),
    ('Организатор', 28),
    ('Мест всего', 12),
    ('Всего заявок', 14),
    ('Подтверждены', 14),
    ('Ожидают', 12),
    ('Лист ожидания', 14),
    ('Посетили', 12),
    ('Не явились', 12),
    ('Отменены', 12),
    ('Заполненность, %', 16),
    ('Создано', 18),
]


def _fill_events_sheet(ws: Worksheet, events) -> None:
    """Заполняет лист «Мероприятия» данными из queryset."""
    start_row = _write_title(
        ws,
        title='Отчёт по мероприятиям',
        subtitle=f'Сформировано: {_russian_now()}',
    )

    for col_idx, (title, _width) in enumerate(EVENTS_COLUMNS, start=1):
        ws.cell(row=start_row, column=col_idx, value=title)
    _apply_header(ws[start_row], ws)
    ws.row_dimensions[start_row].height = 30

    for row_idx, event in enumerate(events, start=start_row + 1):
        total_seats = event.seats_total or 0
        fill_percent = ''
        if total_seats > 0:
            fill_percent = round(
                (event.confirmed_total or 0) * 100 / total_seats, 1
            )
        organizer = ''
        if event.organizer_id:
            organizer = (
                event.organizer.get_full_name() or event.organizer.username
            )
        values = [
            event.pk,
            event.title,
            event.direction.title if event.direction_id else '',
            event.event_type.title if event.event_type_id else '',
            event.get_event_format_display(),
            event.get_status_display(),
            _format_datetime(event.starts_at),
            _format_datetime(event.ends_at),
            event.location,
            organizer,
            total_seats,
            event.registrations_total,
            event.confirmed_total,
            event.pending_total,
            event.waitlist_total,
            event.attended_total,
            event.no_show_total,
            event.cancelled_total,
            fill_percent,
            _format_datetime(event.created_at),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 2:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                cell.alignment = Alignment(vertical='top')

    totals_row = ws.max_row + 1
    if totals_row > start_row + 1:
        ws.cell(row=totals_row, column=1, value='Итого')
        ws.cell(row=totals_row, column=11, value=sum(e.seats_total or 0 for e in events))
        ws.cell(row=totals_row, column=12, value=sum(e.registrations_total for e in events))
        ws.cell(row=totals_row, column=13, value=sum(e.confirmed_total for e in events))
        ws.cell(row=totals_row, column=14, value=sum(e.pending_total for e in events))
        ws.cell(row=totals_row, column=15, value=sum(e.waitlist_total for e in events))
        ws.cell(row=totals_row, column=16, value=sum(e.attended_total for e in events))
        ws.cell(row=totals_row, column=17, value=sum(e.no_show_total for e in events))
        ws.cell(row=totals_row, column=18, value=sum(e.cancelled_total for e in events))
        for cell in ws[totals_row]:
            cell.font = TOTALS_FONT
            cell.fill = TOTALS_FILL

    _autosize_columns(ws, start_row=start_row)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate


def build_events_report() -> Workbook:
    """Формирует workbook с одним листом по мероприятиям."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Мероприятия'

    events = list(
        _annotate_event_stats(
            Event.objects.select_related('direction', 'event_type', 'organizer')
        ).order_by('-starts_at')
    )
    _fill_events_sheet(ws, events)
    return wb


# ---------------------------------------------------------------------------
# Отчёт 2: Регистрации
# ---------------------------------------------------------------------------

REGISTRATIONS_COLUMNS = [
    ('ID', 8),
    ('Мероприятие', 40),
    ('Дата мероприятия', 18),
    ('Участник', 28),
    ('Email', 26),
    ('Телефон', 18),
    ('Организация', 26),
    ('Должность', 22),
    ('Статус', 22),
    ('Источник заявки', 20),
    ('Подана', 18),
    ('Подтверждена', 18),
    ('Отменена', 18),
    ('Причина отмены', 30),
    ('Позиция в ЛО', 12),
    ('Комментарий участника', 30),
]


def _participant_name(registration: EventRegistration) -> str:
    """Возвращает ФИО участника (из снимка заявки или профиля)."""
    if registration.full_name:
        return registration.full_name
    user = registration.user
    full = user.get_full_name() if user else ''
    return full or (user.username if user else '—')


def _fill_registrations_sheet(ws: Worksheet, registrations) -> None:
    """Заполняет лист «Регистрации» данными из queryset."""
    start_row = _write_title(
        ws,
        title='Отчёт по регистрациям',
        subtitle=f'Сформировано: {_russian_now()}',
    )

    for col_idx, (title, _width) in enumerate(REGISTRATIONS_COLUMNS, start=1):
        ws.cell(row=start_row, column=col_idx, value=title)
    _apply_header(ws[start_row], ws)
    ws.row_dimensions[start_row].height = 30

    for row_idx, reg in enumerate(registrations, start=start_row + 1):
        event = reg.event
        user = reg.user
        email = reg.email or (user.email if user else '')
        values = [
            reg.pk,
            event.title if event else '',
            _format_datetime(event.starts_at) if event else '',
            _participant_name(reg),
            email,
            reg.phone or (user.phone if user else ''),
            reg.organization or (user.organization if user else ''),
            reg.position or (user.position if user else ''),
            reg.get_status_display(),
            reg.get_source_display(),
            _format_datetime(reg.created_at),
            _format_datetime(reg.confirmed_at),
            _format_datetime(reg.cancelled_at),
            reg.cancellation_reason or '',
            reg.waitlist_position or '',
            reg.note or '',
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(
                wrap_text=col_idx in {2, 14, 16}, vertical='top'
            )

    _autosize_columns(ws, start_row=start_row)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate


def build_registrations_report() -> Workbook:
    """Формирует workbook с одним листом по регистрациям."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Регистрации'

    registrations = (
        EventRegistration.objects
        .select_related('event', 'user')
        .order_by('-created_at')
    )
    _fill_registrations_sheet(ws, registrations)
    return wb


# ---------------------------------------------------------------------------
# Отчёт 3: Сводный (несколько листов)
# ---------------------------------------------------------------------------

def _fill_summary_sheet(ws: Worksheet) -> None:
    """Заполняет лист «Сводка» общими агрегатами по системе."""
    start_row = _write_title(
        ws,
        title='Сводная статистика',
        subtitle=f'Сформировано: {_russian_now()}',
    )

    events_total = Event.objects.count()
    events_by_status = Counter()
    for status, label in Event.Status.choices:
        events_by_status[label] = Event.objects.filter(status=status).count()

    reg_total = EventRegistration.objects.count()
    reg_by_status = Counter()
    for status, label in EventRegistration.Status.choices:
        reg_by_status[label] = EventRegistration.objects.filter(status=status).count()

    blocks = [
        ('Показатель', 'Значение'),
        ('Всего мероприятий', events_total),
        *[(f'  • {label}', value) for label, value in events_by_status.items()],
        ('', ''),
        ('Всего регистраций', reg_total),
        *[(f'  • {label}', value) for label, value in reg_by_status.items()],
        ('', ''),
        ('Уникальных участников', EventRegistration.objects.values('user_id').distinct().count()),
        ('Количество направлений', Direction.objects.count()),
        ('Активных направлений', Direction.objects.filter(is_active=True).count()),
        ('Количество типов мероприятий', EventType.objects.count()),
        ('Активных типов', EventType.objects.filter(is_active=True).count()),
    ]

    for offset, (label, value) in enumerate(blocks):
        label_cell = ws.cell(row=start_row + offset, column=1, value=label)
        value_cell = ws.cell(row=start_row + offset, column=2, value=value)
        if offset == 0:
            _apply_header([label_cell, value_cell], ws)
            ws.row_dimensions[start_row].height = 26

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 18


def _fill_by_direction_sheet(ws: Worksheet) -> None:
    """Лист «По направлениям»: сколько мероприятий и заявок на каждом."""
    start_row = _write_title(ws, 'Разрез по направлениям')
    headers = ['Направление', 'Мероприятий', 'Всего заявок', 'Подтверждено', 'Ожидают']
    for col_idx, title in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=title)
    _apply_header(ws[start_row], ws)

    rows = (
        Direction.objects.annotate(
            events_total=Count('events', distinct=True),
            reg_total=Count('events__registrations'),
            reg_confirmed=Count(
                'events__registrations',
                filter=Q(events__registrations__status=EventRegistration.Status.CONFIRMED),
            ),
            reg_pending=Count(
                'events__registrations',
                filter=Q(events__registrations__status=EventRegistration.Status.PENDING),
            ),
        )
        .order_by('-reg_total', 'title')
    )
    for offset, direction in enumerate(rows, start=start_row + 1):
        ws.cell(row=offset, column=1, value=direction.title)
        ws.cell(row=offset, column=2, value=direction.events_total)
        ws.cell(row=offset, column=3, value=direction.reg_total)
        ws.cell(row=offset, column=4, value=direction.reg_confirmed)
        ws.cell(row=offset, column=5, value=direction.reg_pending)

    _autosize_columns(ws, start_row=start_row)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate


def _fill_by_type_sheet(ws: Worksheet) -> None:
    """Лист «По типам»: сколько мероприятий и заявок на каждом."""
    start_row = _write_title(ws, 'Разрез по типам мероприятий')
    headers = ['Тип мероприятия', 'Мероприятий', 'Всего заявок', 'Подтверждено']
    for col_idx, title in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=title)
    _apply_header(ws[start_row], ws)

    rows = (
        EventType.objects.annotate(
            events_total=Count('events', distinct=True),
            reg_total=Count('events__registrations'),
            reg_confirmed=Count(
                'events__registrations',
                filter=Q(events__registrations__status=EventRegistration.Status.CONFIRMED),
            ),
        )
        .order_by('-reg_total', 'title')
    )
    for offset, event_type in enumerate(rows, start=start_row + 1):
        ws.cell(row=offset, column=1, value=event_type.title)
        ws.cell(row=offset, column=2, value=event_type.events_total)
        ws.cell(row=offset, column=3, value=event_type.reg_total)
        ws.cell(row=offset, column=4, value=event_type.reg_confirmed)

    _autosize_columns(ws, start_row=start_row)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate


def _fill_top_events_sheet(ws: Worksheet, limit: int = 20) -> None:
    """Лист «Топ мероприятий» по количеству заявок."""
    start_row = _write_title(ws, f'Топ мероприятий по числу заявок (до {limit})')
    headers = [
        'Название',
        'Дата',
        'Направление',
        'Всего заявок',
        'Подтверждено',
        'Ожидают',
        'Лист ожидания',
        'Посетили',
    ]
    for col_idx, title in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=title)
    _apply_header(ws[start_row], ws)

    events = list(
        _annotate_event_stats(
            Event.objects.select_related('direction')
        ).order_by('-registrations_total', '-starts_at')[:limit]
    )
    for offset, event in enumerate(events, start=start_row + 1):
        ws.cell(row=offset, column=1, value=event.title)
        ws.cell(row=offset, column=2, value=_format_datetime(event.starts_at))
        ws.cell(row=offset, column=3, value=event.direction.title if event.direction_id else '')
        ws.cell(row=offset, column=4, value=event.registrations_total)
        ws.cell(row=offset, column=5, value=event.confirmed_total)
        ws.cell(row=offset, column=6, value=event.pending_total)
        ws.cell(row=offset, column=7, value=event.waitlist_total)
        ws.cell(row=offset, column=8, value=event.attended_total)

    _autosize_columns(ws, start_row=start_row)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate


def build_summary_report() -> Workbook:
    """Формирует универсальный отчёт со всеми срезами.

    Листы:

    1. Сводка (общие агрегаты)
    2. Мероприятия (детально)
    3. Регистрации (детально)
    4. По направлениям
    5. По типам
    6. Топ мероприятий
    """
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = 'Сводка'
    _fill_summary_sheet(summary_ws)

    events_ws = wb.create_sheet('Мероприятия')
    events = list(
        _annotate_event_stats(
            Event.objects.select_related('direction', 'event_type', 'organizer')
        ).order_by('-starts_at')
    )
    _fill_events_sheet(events_ws, events)

    registrations_ws = wb.create_sheet('Регистрации')
    registrations = (
        EventRegistration.objects
        .select_related('event', 'user')
        .order_by('-created_at')
    )
    _fill_registrations_sheet(registrations_ws, registrations)

    _fill_by_direction_sheet(wb.create_sheet('По направлениям'))
    _fill_by_type_sheet(wb.create_sheet('По типам'))
    _fill_top_events_sheet(wb.create_sheet('Топ мероприятий'))

    return wb


def build_filename(prefix: str) -> str:
    """Возвращает имя файла отчёта со штампом времени (``prefix_YYYYmmdd_HHMM.xlsx``)."""
    stamp = timezone.localtime().strftime('%Y%m%d_%H%M')
    return f'{prefix}_{stamp}.xlsx'


# Удобный экспорт на случай, если понадобится где-то снаружи.
__all__ = [
    'build_events_report',
    'build_registrations_report',
    'build_summary_report',
    'build_filename',
]
